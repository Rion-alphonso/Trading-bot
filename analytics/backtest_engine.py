import MetaTrader5 as mt5
import pandas as pd
from datetime import datetime, timedelta
import pytz
from utils.logger import system_logger

class BacktestEngine:
    def __init__(self, config):
        self.config = config
        self.symbol = config.get('symbol', 'XAUUSDm')
        self.timeframe_str = config.get('timeframe', 'M5')
        self.tf_map = {
            'M1': mt5.TIMEFRAME_M1, 'M3': mt5.TIMEFRAME_M3, 'M5': mt5.TIMEFRAME_M5, 
            'M15': mt5.TIMEFRAME_M15, 'M30': mt5.TIMEFRAME_M30,
            'H1': mt5.TIMEFRAME_H1, 'H4': mt5.TIMEFRAME_H4, 'D1': mt5.TIMEFRAME_D1
        }
        self.timeframe = self.tf_map.get(self.timeframe_str, mt5.TIMEFRAME_M5)
        
        self.initial_balance = float(config.get('capital', 1110.0))
        self.balance = self.initial_balance
        self.spread = float(config.get('spread', 10.0))
        self.comm = float(config.get('comm', 0.0))
        
        # Risk params
        self.l1_risk = float(config.get('l1_risk', 1.0))
        self.l1_tp = int(config.get('l1_tp', 100))
        self.l1_sl = int(config.get('l1_sl', 1000))
        
        self.l2_risk = float(config.get('l2_risk', 10.0))
        self.l2_tp = int(config.get('l2_tp', 120))
        self.l2_sl = int(config.get('l2_sl', 1000))
        
        self.l3_risk = float(config.get('l3_risk', 100.0))
        self.l3_tp = int(config.get('l3_tp', 150))
        self.l3_sl = int(config.get('l3_sl', 1000))

        # State
        self.strategy = config.get('strategy', 'default')
        self.custom_params = config.get('custom_params', {})
        self.current_level = 1
        
        self.next_direction = 'BUY'
        if self.strategy == 'sell_back':
            self.next_direction = 'SELL'
        elif self.strategy == 'custom' and self.custom_params.get('direction') == 'SELL':
            self.next_direction = 'SELL'
            
        self.trades = []
        self.status = 'ACTIVE'
        self.account_blowups = 0
        self.ist_tz = pytz.timezone('Asia/Kolkata')

    def run(self, start_date=None, end_date=None, num_candles=None, max_trades=None):
        mt5.shutdown()
        mt5_path = r'C:\Program Files\MetaTrader 5\terminal64.exe'
        if not mt5.initialize(path=mt5_path):
            print(f"Failed to initialize MT5 at {mt5_path}. Error: {mt5.last_error()}")
            return False

        all_rates = []
        
        # Calculate a safe start_date to fetch enough data using 30-day chunks
        if num_candles is not None:
            print(f"Starting backtest for {num_candles} candles...")
            # Approximate days needed (M5 = 288 candles per day max, let's use 200 to be safe)
            days_needed = int(num_candles / 200) + 2
            start_date = end_date - timedelta(days=days_needed)
        elif max_trades is not None:
            print(f"Starting backtest for exactly {max_trades} trades...")
            # Very conservative estimate: 1 trade per day
            start_date = end_date - timedelta(days=max_trades)
        else:
            print(f"Starting backtest from {start_date} to {end_date}...")

        # Unified 30-day chunk fetching to bypass all MT5 limits
        chunk_end = end_date
        while chunk_end > start_date:
            chunk_start = max(start_date, chunk_end - timedelta(days=30))
            rates = mt5.copy_rates_range(self.symbol, self.timeframe, chunk_start, chunk_end)
            
            if rates is not None and len(rates) > 0:
                all_rates.append(pd.DataFrame(rates))
            
            chunk_end = chunk_start
            
        if not all_rates:
            print(f"No data retrieved. MT5 error: {mt5.last_error()}")
            return False
            
        df = pd.concat(all_rates[::-1]).drop_duplicates('time').reset_index(drop=True)
        
        # If user asked for exact candles, slice the tail
        if num_candles is not None and len(df) > num_candles:
            df = df.tail(num_candles).reset_index(drop=True)
        df['time'] = pd.to_datetime(df['time'], unit='s')
        
        # Store OHLC for frontend lightweight charts
        self.rates_data = df[['time', 'open', 'high', 'low', 'close', 'tick_volume']].copy()
        self.rates_data.rename(columns={'tick_volume': 'volume'}, inplace=True)
        self.rates_data['time'] = self.rates_data['time'].dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        
        symbol_info = mt5.symbol_info(self.symbol)
        if not symbol_info:
            print("Symbol info not found.")
            return False
            
        point = symbol_info.point
        contract_size = symbol_info.trade_contract_size

        print(f"Retrieved {len(df)} candles. Simulating trades...")

        active_trade = None
        
        for index, row in df.iterrows():
            if self.status == 'STOPPED':
                break
                
            if max_trades and len(self.trades) >= max_trades and not active_trade:
                # We have hit the target trade count and have no open positions. 
                # Trim the chart data so it doesn't show years of blank space.
                self.rates_data = self.rates_data.iloc[:index+1].copy()
                break
                
            candle_time = row['time']
            open_p, high_p, low_p, close_p = row['open'], row['high'], row['low'], row['close']
            
            # If no active trade, check session and open
            if not active_trade:
                if self.current_level == 1 and not self._is_active_session(candle_time):
                    pass
                else:
                    volume, sl_price, tp_price = self._calculate_params(open_p, point, contract_size)
                    if volume > 0:
                        active_trade = {
                            'entry_time': candle_time,
                            'direction': self.next_direction,
                            'entry_price': open_p,
                            'sl_price': sl_price,
                            'tp_price': tp_price,
                            'level': self.current_level,
                            'volume': volume
                        }

            # If we have an active trade, check if it hit SL or TP in this candle
            if active_trade:
                hit_result = self._check_trade_closure(active_trade, high_p, low_p)
                if hit_result:
                    # Close trade
                    exit_price = active_trade['tp_price'] if hit_result == 'TP' else active_trade['sl_price']
                    
                    price_diff = exit_price - active_trade['entry_price']
                    if active_trade['direction'] == 'SELL':
                        price_diff = -price_diff
                        
                    profit = price_diff * active_trade['volume'] * contract_size
                    
                    # Apply spread and commission penalty
                    spread_cost = self.spread * point * active_trade['volume'] * contract_size
                    comm_cost = self.comm * active_trade['volume']
                    profit = profit - spread_cost - comm_cost
                    
                    self.balance += profit
                    
                    active_trade['exit_time'] = candle_time
                    active_trade['exit_price'] = exit_price
                    active_trade['profit'] = profit
                    active_trade['result'] = hit_result
                    active_trade['balance_after'] = self.balance
                    
                    self.trades.append(active_trade)
                    
                    if self.balance <= 0:
                        self.account_blowups += 1
                        print(f"Account blown up at {candle_time}! Stopping backtest.")
                        self.status = 'STOPPED'
                        active_trade = None
                        break
                    
                    # Process Martingale
                    # Always stay at Level 1 per new strategy rules
                    self.current_level = 1
                    
                    # Toggle direction
                    if self.strategy == 'buy_back':
                        self.next_direction = 'BUY'
                    elif self.strategy == 'sell_back':
                        self.next_direction = 'SELL'
                    elif self.strategy == 'custom':
                        custom_dir = self.custom_params.get('direction', 'BOTH')
                        if custom_dir == 'BUY':
                            self.next_direction = 'BUY'
                        elif custom_dir == 'SELL':
                            self.next_direction = 'SELL'
                        else:
                            self.next_direction = 'SELL' if self.next_direction == 'BUY' else 'BUY'
                    else:
                        self.next_direction = 'SELL' if self.next_direction == 'BUY' else 'BUY'
                    active_trade = None


        # Force-close any open trade at the end of the backtest
        if active_trade and not df.empty:
            last_candle = df.iloc[-1]
            exit_price = last_candle['close']
            
            price_diff = exit_price - active_trade['entry_price']
            if active_trade['direction'] == 'SELL':
                price_diff = -price_diff
                
            profit = price_diff * active_trade['volume'] * contract_size
            spread_cost = self.spread * point * active_trade['volume'] * contract_size
            comm_cost = self.comm * active_trade['volume']
            profit = profit - spread_cost - comm_cost
            
            self.balance += profit
            
            active_trade['exit_time'] = last_candle['time']
            active_trade['exit_price'] = exit_price
            active_trade['profit'] = profit
            active_trade['result'] = 'END_OF_TEST'
            active_trade['balance_after'] = self.balance
            
            self.trades.append(active_trade)
            
            if self.balance <= 0 and self.account_blowups == 0:
                self.account_blowups = 1

        self._generate_excel()
        return True

    def _check_trade_closure(self, trade, high, low):
        # Conservative check: If both SL and TP are within the candle bounds, assume SL hit first.
        if trade['direction'] == 'BUY':
            if low <= trade['sl_price']: return 'SL'
            if high >= trade['tp_price']: return 'TP'
        else: # SELL
            if high >= trade['sl_price']: return 'SL'
            if low <= trade['tp_price']: return 'TP'
        return None

    def _is_active_session(self, utc_time):
        if self.strategy == '24_7':
            return True
            
        if hasattr(utc_time, 'tzinfo') and utc_time.tzinfo is not None:
            naive_time = utc_time.replace(tzinfo=None)
        else:
            naive_time = utc_time
        ist_time = naive_time.replace(tzinfo=pytz.utc).astimezone(self.ist_tz).time()
        
        if self.strategy == 'custom':
            sess_str = self.custom_params.get('sessions', '')
            if not sess_str.strip():
                return True # empty means 24h
            for s in sess_str.split(','):
                try:
                    start_str, end_str = s.split('-')
                    start_t = datetime.strptime(start_str.strip(), "%H:%M").time()
                    end_t = datetime.strptime(end_str.strip(), "%H:%M").time()
                    if start_t <= ist_time <= end_t:
                        return True
                except:
                    pass
            return False

        # Default behavior: 10:00 - 17:30 and 20:00 - 22:30
        if datetime.strptime("10:00", "%H:%M").time() <= ist_time <= datetime.strptime("17:30", "%H:%M").time():
            return True
        if datetime.strptime("20:00", "%H:%M").time() <= ist_time <= datetime.strptime("22:30", "%H:%M").time():
            return True
        return False

    def _calculate_params(self, entry_price, point, contract_size):
        # Risk adjustment
        if self.strategy == 'advanced_grow':
            risk_pct = 1.0
        elif self.strategy == 'custom':
            risk_pct = float(self.custom_params.get('risk', 0.1))
        else:
            # Use configured risk based on martingale level instead of hardcoded 0.1
            risk_pct = float(self.l1_risk) if self.current_level == 1 else (float(self.l2_risk) if self.current_level == 2 else float(self.l3_risk))

        sl_pts = self.l1_sl if self.current_level == 1 else (self.l2_sl if self.current_level == 2 else self.l3_sl)
        tp_pts = self.l1_tp if self.current_level == 1 else (self.l2_tp if self.current_level == 2 else self.l3_tp)

        if self.strategy == 'custom':
            sl_pts = int(self.custom_params.get('sl', 1000))
            tp_pts = int(self.custom_params.get('tp', 10000))
            
        risk_amount = self.balance * (risk_pct / 100.0)
        
        # Total cost of losing 1 lot (SL movement + spread cost + commission)
        loss_per_lot = (sl_pts + self.spread) * point * contract_size + self.comm
        
        if loss_per_lot > 0:
            volume = round(risk_amount / loss_per_lot, 2)
        else:
            volume = 0.01
            
        if volume < 0.01:
            volume = 0.01
            
        sl_price = entry_price - (sl_pts * point) if self.next_direction == 'BUY' else entry_price + (sl_pts * point)
        tp_price = entry_price + (tp_pts * point) if self.next_direction == 'BUY' else entry_price - (tp_pts * point)
        
        return volume, sl_price, tp_price

    def _generate_excel(self):
        print("Generating Excel Report...")
        if not self.trades:
            print("No trades taken in this period.")
            return

        df = pd.DataFrame(self.trades)
        
        # Convert to IST safely (handle both tz-naive and tz-aware timestamps)
        for col in ['entry_time', 'exit_time']:
            df[col] = pd.to_datetime(df[col])
            if df[col].dt.tz is None:
                df[col] = df[col].dt.tz_localize('UTC')
            df[col] = df[col].dt.tz_convert(self.ist_tz).dt.tz_localize(None)
        
        wins = df[df['result'] == 'TP']
        losses = df[df['result'] == 'SL']
        
        total_trades = len(df)
        win_rate = (len(wins) / total_trades) * 100 if total_trades > 0 else 0
        total_profit = df['profit'].sum()
        
        l2_activations = len(df[df['level'] == 2])
        l3_activations = len(df[df['level'] == 3])

        dash_data = {
            "Metric": [
                "Initial Capital", "Final Capital", "Total Profit",
                "Total Trades", "Wins (TP)", "Losses (SL)", "Win Rate %",
                "Level 2 Activations", "Level 3 Activations", "Account Blowups"
            ],
            "Value": [
                self.initial_balance, self.balance, total_profit,
                total_trades, len(wins), len(losses), round(win_rate, 2),
                l2_activations, l3_activations, self.account_blowups
            ]
        }
        
        dash_df = pd.DataFrame(dash_data)

        filename = f"Backtest_Results_{self.strategy}.xlsx"
        from openpyxl.styles import PatternFill
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            dash_df.to_excel(writer, sheet_name='Dashboard', index=False)
            df.to_excel(writer, sheet_name='Trade Log', index=False)
            
            # Formatting
            workbook = writer.book
            worksheet = workbook['Trade Log']
            
            # Yellow fill for Level 2, Red fill for Level 3
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            try:
                level_col_idx = list(df.columns).index('level') + 1
                for row_idx in range(2, worksheet.max_row + 1):
                    level_val = worksheet.cell(row=row_idx, column=level_col_idx).value
                    if level_val == 2:
                        for cell in worksheet[row_idx]:
                            cell.fill = yellow_fill
                    elif level_val == 3:
                        for cell in worksheet[row_idx]:
                            cell.fill = red_fill
            except ValueError:
                pass # 'level' column not found
        
        print(f"Done! Saved to {filename}")
